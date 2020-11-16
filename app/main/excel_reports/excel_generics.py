from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Font, Color, colors
from django.utils import timezone

def as_text(value):
    if value is None:
        return ""
    return str(value)


def merge_cell(worksheet, row_num=1, coll_num=1, bolt=False, vertical=False, merge_num=1):
    if vertical:
        end_merge_num = coll_num + merge_num - 1
        # print(f'start_row={row_num}, start_column={coll_num}, end_row={row_num}, end_column={end_merge_num}')
        worksheet.merge_cells(start_row=row_num, start_column=coll_num, end_row=row_num,
                              end_column=end_merge_num)
        for i in range(1, merge_num + 1):
            cell = worksheet.cell(row=row_num, column=end_merge_num)
            set_appearance(cell, bolt=bolt)
    else:
        end_merge_num = row_num + merge_num
        worksheet.merge_cells(start_row=row_num, start_column=coll_num, end_row=end_merge_num,
                              end_column=coll_num)
        for i in range(1, merge_num + 1):
            cell = worksheet.cell(row=end_merge_num, column=coll_num)
            set_appearance(cell, bolt=bolt)
    return worksheet


def merge_cell_new(worksheet, row_num=1, coll_num=1, bolt=False, vertical=False, merge_num=1):
    if vertical:
        end_merge_num = coll_num + merge_num - 1
        # print(f'start_row={row_num}, start_column={coll_num}, end_row={row_num}, end_column={end_merge_num}')
        worksheet.merge_cells(start_row=row_num, start_column=coll_num, end_row=row_num,
                              end_column=end_merge_num)
        for i in range(1, merge_num + 1):
            cell = worksheet.cell(row=row_num, column=end_merge_num)
            set_appearance(cell, bolt=bolt)
    else:
        end_merge_num = row_num + merge_num
        worksheet.merge_cells(start_row=row_num, start_column=coll_num, end_row=end_merge_num,
                              end_column=coll_num)
        for i in range(1, merge_num + 1):
            cell = worksheet.cell(row=end_merge_num, column=coll_num)
            set_appearance(cell, bolt=bolt)
    return worksheet


def expands_columns_width(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length
    return worksheet


def set_appearance(table_cell, border=True, bolt=True, horizontal_alignment='center', u=None, color=colors.BLACK):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    table_cell.font = Font(bold=bolt, u=u, color=color)
    table_cell.alignment = Alignment(horizontal=horizontal_alignment)
    if border:
        table_cell.border = thin_border


def insert_data_to_cell(worksheet, cell_value, row_num=1, coll_num=1, bolt=False,
                        merge=False, merge_num=1, vertical=True):
    if merge:
        merge_cell(worksheet, row_num=row_num, coll_num=coll_num, bolt=bolt, vertical=vertical, merge_num=merge_num)
    cell = worksheet.cell(row=row_num, column=coll_num)
    cell.value = cell_value
    set_appearance(cell, bolt=bolt)


def add_data_to_cell(worksheet, cell_value, row_num=1, coll_num=1, bolt=False,
                        merge=False, merge_num=1, vertical=True):
    if vertical:
        row_num, coll_num = row_num, coll_num
    else:
        row_num, coll_num = coll_num, row_num

    if merge:
        merge_cell(worksheet, row_num=row_num, coll_num=coll_num, bolt=bolt, vertical=vertical, merge_num=merge_num)
    cell = worksheet.cell(row=row_num, column=coll_num)
    cell.value = cell_value
    set_appearance(cell, bolt=bolt)


def cell_data_generate(worksheet, value, row_num, coll_num, bolt=False):
    cell = worksheet.cell(row=row_num, column=coll_num)
    cell.value = value
    set_appearance(cell, bolt=bolt)


def step_enumerate(input_list, start=0, step=1):
    for elem in input_list:
        yield (start, elem)
        start += step


# def add_table_titles(worksheet, titles, start_position=1, vertical=True, bolt=True):
#     for position in titles.keys():
#         if vertical:
#             insert_data_to_cell(worksheet, titles[position], position, start_position, bolt=bolt)
#         else:
#             insert_data_to_cell(worksheet, titles[position], start_position, position, bolt=bolt)


def add_table_titles_new(worksheet, titles, start_position=1, vertical=True, bolt=True):
    for position in titles.keys():
        if vertical:
            insert_data_to_cell(worksheet, titles[position]['title'], position, start_position, bolt=bolt)
        else:
            insert_data_to_cell(worksheet, titles[position]['title'], start_position, position, bolt=bolt)


def insert_table_titles(worksheet, titles, start_position=1, vertical=True, bolt=True):
    for position, title in enumerate(titles.keys(), 1):
        if vertical:
            insert_data_to_cell(worksheet, title, position, start_position, bolt=bolt, vertical=vertical)
        else:
            insert_data_to_cell(worksheet, title, start_position, position, bolt=bolt, vertical=vertical)


# def add_table_data(worksheet, fields, vertical=True, start_position=1):
#     for position in fields.keys():
#         print(position)
#         if vertical:
#             insert_data_to_cell(worksheet, fields[position], position, start_position)
#         else:
#             insert_data_to_cell(worksheet, fields[position], start_position, position)


def get_net_compile_id(cell, report_id, display='ссылка'):
    url = f'http://172.16.25.10/netcompiles/{id}'
    cell.value = f'=HYPERLINK({url}, {display})'
    cell.font = Font(u='single', color=colors.BLUE)


def insert_url_to_cell(worksheet, url_string, url_name='ссылка', row_num=1, coll_num=1, bolt=False,
                        merge=False, merge_num=1, vertical=True):
    if merge:
        merge_cell(worksheet, row_num=row_num, coll_num=coll_num, bolt=bolt, vertical=vertical, merge_num=merge_num)
    cell = worksheet.cell(row=row_num, column=coll_num)
    cell.hyperlink = url_string
    cell.value = url_name
    set_appearance(cell, bolt=bolt, u='single', color=colors.BLUE)


def get_grafana_url(object):
    dashboard_url = object.chaos.grafana_dashboard_url
    create_date_time = str(timezone.localtime(object.create_date_time).strftime('%s')) + '000'
    if object.date_time_finish is None:
        date_time_finish = 'now'
    else:
        date_time_finish = str(timezone.localtime(object.date_time_finish).strftime('%s')) + '000'
    href = f"{dashboard_url}?orgId=1&from={create_date_time}&to={date_time_finish}"
    # href = f"{ object.chaos.grafana_dashboard_url }{ object.create_date_time|date:"U" }000&to={% if object.date_time_finish%}{{ object.date_time_finish|date:"U" }}000{% else %}now{% endif %}"
    print(href)
    return href
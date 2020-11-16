import django
import os
from functools import partial
from main.models import DrawImgsStat, NetCompilationStat, NetCompileReport, DrawImgsReport
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment
from main.excel_reports.excel_generics import expands_columns_width, step_enumerate, \
    insert_table_titles, add_data_to_cell, insert_url_to_cell, get_grafana_url

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "conf.settings")
django.setup()


def get_columns_ids(model_included_columns, reports_list):
    columns = []
    columns_ids = []
    db_model_fields = [field.name for field in reports_list.model._meta.fields]

    for num, field in enumerate(model_included_columns):
        if field in db_model_fields:
            columns.append(reports_list.model._meta.get_field(field).verbose_name)
            columns_ids.append(db_model_fields.index(field))
    return columns, columns_ids


def generate_table_data(worksheet, reports_list, model_included_columns, vertical=False,
                        start_row_num=1, alignment='center'):
    columns_and_ids = get_columns_ids(model_included_columns, reports_list)
    columns = columns_and_ids[0]
    columns_ids = columns_and_ids[1]
    colon_or_row_num = 1
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # создание шапки таблицы
    for elem_num, column_title in enumerate(columns, start_row_num):
        if vertical:
            cell = worksheet.cell(row=elem_num, column=colon_or_row_num)
        else:
            cell = worksheet.cell(row=colon_or_row_num, column=elem_num)
        cell.value = column_title
        cell.border = thin_border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # заполнение таблицы даными
    for stat in reports_list.values_list():
        colon_or_row_num += 1
        row = [stat[stat_id] for stat_id in columns_ids]
        for elem_num, cell_value in enumerate(row, start_row_num):
            if vertical:
                cell = worksheet.cell(row=elem_num, column=colon_or_row_num)
            else:
                cell = worksheet.cell(row=colon_or_row_num, column=elem_num)
            cell.value = cell_value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal=alignment)
    return worksheet


def create_excel_cheet(workbook, reports_list, model_included_columns, vertical=False, alignment='center'):
    worksheet = workbook.create_sheet(reports_list.model._meta.verbose_name_plural.title(), 0)
    worksheet_with_data = generate_table_data(worksheet, reports_list, model_included_columns,
                                              vertical=vertical, alignment=alignment)
    # увеличение ширины колонок
    expands_columns_width(worksheet_with_data)
    return workbook


def create_excel_cheet_for_stats(workbook, reports_list, report_stats_model, print_columns, vertical=False):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for report in reports_list:
        if report_stats_model == DrawImgsStat:
            statistic_objects = DrawImgsStat.objects.filter(draw_imgs_report=report)
        if report_stats_model == NetCompilationStat:
            statistic_objects = NetCompilationStat.objects.filter(net_compile_report=report)
        row_num = 1
        colon_num = 1
        worksheet = workbook.create_sheet(report.__str__(), 0)
        columns = []
        columns_ids = []
        db_model_fields = report_stats_model._meta.fields

        for num, field in enumerate(db_model_fields):
            if field.name in print_columns:
                columns.append(report_stats_model._meta.get_field(field.name).verbose_name)
                columns_ids.append(num)

        # создание шапки таблицы
        for elem_num, column_title in enumerate(columns, 1):
            if vertical:
                cell = worksheet.cell(row=elem_num, column=colon_num)
            else:
                cell = worksheet.cell(row=row_num, column=elem_num)
            cell.value = column_title
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # заполнение таблицы даными
        for statistic_object in statistic_objects.values_list():
            row_num += 1
            colon_num += 1
            rows_list = [statistic_object[stat_id] for stat_id in columns_ids]
            for elem_num, cell_value in enumerate(rows_list, 1):
                if vertical:
                    cell = worksheet.cell(row=elem_num, column=colon_num)
                else:
                    cell = worksheet.cell(row=row_num, column=elem_num)
                cell.value = cell_value
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # увеличение ширины колонок
        expands_columns_width(worksheet)
    return workbook


def create_draw_imgs_stats_sheet(workbook, report, vertical=False, start_position=1):
    report_statistics = DrawImgsStat.objects.filter(draw_imgs_report=report)
    worksheet = workbook.create_sheet(report.__str__(), 0)

    sheet_datas = {'Процент отрисовки': partial(add_data_to_cell),
                   'Затраченное время': partial(add_data_to_cell),
                   'Отрисованных ценников': partial(add_data_to_cell),
                   }
    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    start_position += 1
    for elem_num, statistic in enumerate(report_statistics, start_position):
        sheet_datas['Процент отрисовки'](worksheet, statistic.percent_step, 1, elem_num, vertical=vertical)
        sheet_datas['Затраченное время'](worksheet, statistic.elapsed_time, 2, elem_num, vertical=vertical)
        sheet_datas['Отрисованных ценников'](worksheet, statistic.images_succeeded, 3, elem_num, vertical=vertical)

    # увеличение ширины колонок
    expands_columns_width(worksheet)
    return workbook


def create_draw_imgs_sheet(workbook, report, vertical=False, start_position=1):
    worksheet = workbook.create_sheet('Отчет об отрисовке', 0)

    sheet_datas = {'Дата и время создания отчета': partial(add_data_to_cell),
                   'Статус ': partial(add_data_to_cell),
                   'Дата и время окончания отчета': partial(add_data_to_cell),
                   'Время отрисовки 50%': partial(add_data_to_cell),
                   'Время отрисовки 75%': partial(add_data_to_cell),
                   'Время отрисовки 90%': partial(add_data_to_cell),
                   'Время отрисовки 95%': partial(add_data_to_cell),
                   'Время отрисовки 96%': partial(add_data_to_cell),
                   'Время отрисовки 97%': partial(add_data_to_cell),
                   'Время отрисовки 98%': partial(add_data_to_cell),
                   'Время отрисовки 99%': partial(add_data_to_cell),
                   'Время отрисовки 99.5%': partial(add_data_to_cell),
                   'Время отрисовки 99.9%': partial(add_data_to_cell),
                   'Время отрисовки 100%': partial(add_data_to_cell),
                   'Среднее значение потребления в процессе отрисовки, mA': partial(add_data_to_cell),
                   }
    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    start_position += 1
    sheet_datas['Дата и время создания отчета'](worksheet, report.create_date_time, 1, start_position, vertical=vertical)
    sheet_datas['Статус '](worksheet, report.status, 2, start_position, vertical=vertical)
    sheet_datas['Дата и время окончания отчета'](worksheet, report.date_time_finish, 3, start_position,
                                                 vertical=vertical)
    sheet_datas['Время отрисовки 50%'](worksheet, report.p50, 4, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 75%'](worksheet, report.p75, 5, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 90%'](worksheet, report.p90, 6, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 95%'](worksheet, report.p95, 7, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 96%'](worksheet, report.p96, 8, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 97%'](worksheet, report.p97, 9, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 98%'](worksheet, report.p98, 10, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 99%'](worksheet, report.p99, 11, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 99.5%'](worksheet, report.p995, 12, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 99.9%'](worksheet, report.p999, 13, start_position, vertical=vertical)
    sheet_datas['Время отрисовки 100%'](worksheet, report.p100, 14, start_position, vertical=vertical)
    sheet_datas['Среднее значение потребления в процессе отрисовки, mA'](worksheet, report.voltage_average, 15,
                                                                         start_position,
                                                                         vertical=vertical)

    # увеличение ширины колонок
    expands_columns_width(worksheet)
    return workbook


def create_net_compile_stats_sheet(workbook, report, vertical=False, start_position=1):
    report_statistics = NetCompilationStat.objects.filter(net_compile_report=report)
    worksheet = workbook.create_sheet(report.__str__(), 0)

    sheet_datas = {'Ценников онлайн': partial(add_data_to_cell),
                   'Компиляция сети (%)': partial(add_data_to_cell),
                   'Затраченное время': partial(add_data_to_cell),
                   }

    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    # заполнение таблицы даными
    start_position += 1
    for elem_num, statistic in enumerate(report_statistics, start_position):
        sheet_datas['Ценников онлайн'](worksheet, statistic.online_esl, 1, elem_num, vertical=vertical)
        sheet_datas['Компиляция сети (%)'](worksheet, statistic.compilation_percent, 2, elem_num, vertical=vertical)
        sheet_datas['Затраченное время'](worksheet, statistic.elapsed_time, 3, elem_num, vertical=vertical)

    # увеличение ширины колонок
    expands_columns_width(worksheet)
    return workbook


def create_net_compile_sheet(workbook, report, vertical=False, start_position=1):
    worksheet = workbook.create_sheet('Отчет о сборке сети', 0)

    sheet_datas = {'Дата и время начала сборки': partial(add_data_to_cell),
                   'Статус': partial(add_data_to_cell),
                   'Дата и время окончания сборки': partial(add_data_to_cell),
                   'Время сборки 50%': partial(add_data_to_cell),
                   'Время сборки 75%': partial(add_data_to_cell),
                   'Время сборки 90%': partial(add_data_to_cell),
                   'Время сборки 95%': partial(add_data_to_cell),
                   'Время сборки 96%': partial(add_data_to_cell),
                   'Время сборки 97%': partial(add_data_to_cell),
                   'Время сборки 98%': partial(add_data_to_cell),
                   'Время сборки 99%': partial(add_data_to_cell),
                   'Время сборки 99.5%': partial(add_data_to_cell),
                   'Время сборки 99.9%': partial(add_data_to_cell),
                   'Время сборки 100%': partial(add_data_to_cell),
                   'Среднее значение потребления в процессе сборки, mA': partial(add_data_to_cell),
                   }
    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    # заполнение таблицы даными
    start_position += 1
    sheet_datas['Дата и время начала сборки'](worksheet, report.create_date_time, 1, start_position, merge=True,
                                              vertical=vertical)
    sheet_datas['Статус'](worksheet, report.status, 2, start_position, merge=True, vertical=vertical)
    sheet_datas['Дата и время окончания сборки'](worksheet, report.date_time_finish, 3, start_position, merge=True,
                                                 vertical=vertical)
    sheet_datas['Время сборки 50%'](worksheet, report.p50, 4, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 75%'](worksheet, report.p75, 5, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 90%'](worksheet, report.p90, 6, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 95%'](worksheet, report.p95, 7, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 96%'](worksheet, report.p96, 8, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 97%'](worksheet, report.p97, 9, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 98%'](worksheet, report.p98, 10, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 99%'](worksheet, report.p99, 11, start_position, merge=True,
                                    vertical=vertical)
    sheet_datas['Время сборки 99.5%'](worksheet, report.p995, 12, start_position, merge=True,
                                      vertical=vertical)
    sheet_datas['Время сборки 99.9%'](worksheet, report.p999, 13, start_position, merge=True,
                                      vertical=vertical)
    sheet_datas['Время сборки 100%'](worksheet, report.p100, 14, start_position, merge=True,
                                     vertical=vertical)
    sheet_datas['Среднее значение потребления в процессе сборки, mA'](worksheet, report.voltage_average, 15,
                                                                      start_position,
                                                                      merge=True,
                                                                      vertical=vertical)

    # увеличение ширины колонок
    expands_columns_width(worksheet)
    return workbook


def create_all_statistics_sheet(workbook, report, vertical=False, start_position=1):
    worksheet = workbook.create_sheet('Общая статистика', 0)

    sheet_datas = {'Дата и время': partial(add_data_to_cell),
                   'total_nodes': partial(add_data_to_cell),
                   'inaccessible_nodes': partial(add_data_to_cell),
                   'total_number_routes': partial(add_data_to_cell),
                   'maximum_road_length': partial(add_data_to_cell),
                   'average_route_length': partial(add_data_to_cell),
                   'accessible_nodes_percent': partial(add_data_to_cell),
                   'elapsed_time': partial(add_data_to_cell),
                   'total_esl': partial(add_data_to_cell),
                   'online_esl': partial(add_data_to_cell),
                   'images_in_transit': partial(add_data_to_cell),
                   'images_in_draw': partial(add_data_to_cell),
                   'images_in_resend_queue': partial(add_data_to_cell),
                   'images_succeeded': partial(add_data_to_cell),
                   'images_failed': partial(add_data_to_cell),
                   'currently_scanning': partial(add_data_to_cell),
                   'network_mode': partial(add_data_to_cell),
                   'connects': partial(add_data_to_cell),
                   'network_mode_percent': partial(add_data_to_cell),
                   'voltage_average': partial(add_data_to_cell),
                   'voltage_current': partial(add_data_to_cell),
                   'voltage_max': partial(add_data_to_cell),
                   'bat_reserved1': partial(add_data_to_cell),
                   'bat_reserved2': partial(add_data_to_cell),
                   'bat_reserved3': partial(add_data_to_cell),
                   'bat_reserved4': partial(add_data_to_cell),
                   }
    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    # заполнение таблицы даными
    start_position += 1
    for elem_num, statistic in enumerate(report, start_position):
        sheet_datas['Дата и время'](worksheet, statistic.date_time, 1, elem_num, vertical=vertical)
        sheet_datas['total_nodes'](worksheet, statistic.total_nodes, 2, elem_num, vertical=vertical)
        sheet_datas['inaccessible_nodes'](worksheet, statistic.inaccessible_nodes, 3, elem_num, vertical=vertical)
        sheet_datas['total_number_routes'](worksheet, statistic.total_number_routes, 4, elem_num, vertical=vertical)
        sheet_datas['maximum_road_length'](worksheet, statistic.maximum_road_length, 5, elem_num, vertical=vertical)
        sheet_datas['average_route_length'](worksheet, statistic.average_route_length, 6, elem_num, vertical=vertical)
        sheet_datas['accessible_nodes_percent'](worksheet, statistic.accessible_nodes_percent, 7, elem_num,
                                                vertical=vertical)
        sheet_datas['elapsed_time'](worksheet, statistic.elapsed_time, 8, elem_num, vertical=vertical)
        sheet_datas['total_esl'](worksheet, statistic.total_esl, 9, elem_num, vertical=vertical)
        sheet_datas['online_esl'](worksheet, statistic.online_esl, 10, elem_num, vertical=vertical)
        sheet_datas['images_in_transit'](worksheet, statistic.images_in_transit, 11, elem_num, vertical=vertical)
        sheet_datas['images_in_draw'](worksheet, statistic.images_in_draw, 12, elem_num, vertical=vertical)
        sheet_datas['images_in_resend_queue'](worksheet, statistic.images_in_resend_queue, 13, elem_num,
                                              vertical=vertical)
        sheet_datas['images_succeeded'](worksheet, statistic.images_succeeded, 14, elem_num, vertical=vertical)
        sheet_datas['images_failed'](worksheet, statistic.images_failed, 15, elem_num, vertical=vertical)
        sheet_datas['currently_scanning'](worksheet, statistic.currently_scanning, 16, elem_num, vertical=vertical)
        sheet_datas['network_mode'](worksheet, statistic.network_mode, 17, elem_num, vertical=vertical)
        sheet_datas['connects'](worksheet, statistic.connects, 18, elem_num, vertical=vertical)
        sheet_datas['network_mode_percent'](worksheet, statistic.network_mode_percent, 19, elem_num, vertical=vertical)
        sheet_datas['voltage_average'](worksheet, statistic.voltage_average, 20, elem_num, vertical=vertical)
        sheet_datas['voltage_current'](worksheet, statistic.voltage_current, 21, elem_num, vertical=vertical)
        sheet_datas['voltage_max'](worksheet, statistic.voltage_max, 22, elem_num, vertical=vertical)
        sheet_datas['bat_reserved1'](worksheet, statistic.bat_reserved1, 23, elem_num, vertical=vertical)
        sheet_datas['bat_reserved2'](worksheet, statistic.bat_reserved2, 24, elem_num, vertical=vertical)
        sheet_datas['bat_reserved3'](worksheet, statistic.bat_reserved3, 25, elem_num, vertical=vertical)
        sheet_datas['bat_reserved4'](worksheet, statistic.bat_reserved4, 26, elem_num, vertical=vertical)

    # увеличение ширины колонок
    expands_columns_width(worksheet)
    return workbook


def create_common_sheet(workbook, metric_report, server_ip, vertical=False, start_position=1):
    net_reports = NetCompileReport.objects.filter(metric_report=metric_report).order_by('create_date_time')
    draw_reports = DrawImgsReport.objects.filter(metric_report=metric_report).order_by('create_date_time')
    net_report_counter = 0
    draw_report_counter = 0
    worksheet = workbook.create_sheet('Сводный отчет', 0)


    sheet_datas = {'  ': partial(insert_url_to_cell),
                   'Графики этапа сборки сети': partial(insert_url_to_cell),
                   'Дата и время начала сборки': partial(add_data_to_cell),
                   'Время сборки сети': partial(add_data_to_cell),
                   'Процент сборки сети': partial(add_data_to_cell),
                   'Собранность сети за 60 мин.,%': partial(add_data_to_cell),
                   'Фактическое количество ESL': partial(add_data_to_cell),
                   'Статус': partial(add_data_to_cell),
                   'Среднее значение потребления в процессе сборки, mA': partial(add_data_to_cell),
                   ' ': partial(insert_url_to_cell),
                   'Графики этапа отрисовки ценников': partial(insert_url_to_cell),
                   'Время отрисовки 50%': partial(add_data_to_cell),
                   'Время отрисовки 75%': partial(add_data_to_cell),
                   'Время отрисовки 90%': partial(add_data_to_cell),
                   'Время отрисовки 95%': partial(add_data_to_cell),
                   'Время отрисовки 96%': partial(add_data_to_cell),
                   'Время отрисовки 97%': partial(add_data_to_cell),
                   'Время отрисовки 98%': partial(add_data_to_cell),
                   'Время отрисовки 99%': partial(add_data_to_cell),
                   'Время отрисовки 99.5%': partial(add_data_to_cell),
                   'Время отрисовки 99.9%': partial(add_data_to_cell),
                   'Время отрисовки 100%': partial(add_data_to_cell),
                   'Не отрисовано, шт ': partial(add_data_to_cell),
                   'Среднее значение потребления в процессе отрисовки, mA': partial(add_data_to_cell),
                   'Средний размер изображений ценников': partial(add_data_to_cell),
                   'Среднее значение потребления (общее), mA': partial(add_data_to_cell),
                   }

    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    # Расширение колонок
    expands_columns_width(worksheet)

    # заполнение таблицы даными
    start_position += 1
    merge_cell_num = metric_report.draw_imgs_amount
    for elem_num, report in step_enumerate(net_reports, start_position, merge_cell_num):
        net_report_counter += 1
        sheet_datas['  '](worksheet, f'http://{server_ip}/netcompiles/{report.pk}',
                          f'Отчет о сборке #{net_report_counter}({report.pk})',
                          1, elem_num, merge=True, merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Графики этапа сборки сети'](worksheet, get_grafana_url(report), f'смотреть',
                          2, elem_num, merge=True, merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Дата и время начала сборки'](worksheet, report.create_date_time, 3, elem_num, merge=True,
                                                  merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Время сборки сети'](worksheet, report.elapsed_time, 4, elem_num, merge=True,
                                         merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Процент сборки сети'](worksheet, report.final_percent, 5, elem_num, merge=True,
                                           merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Собранность сети за 60 мин.,%'](worksheet, report.t60, 6, elem_num, merge=True,
                                                     merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Фактическое количество ESL'](worksheet, report.fact_total_esl, 7, elem_num, merge=True,
                                                  merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Статус'](worksheet, report.status, 8, elem_num, merge=True, merge_num=merge_cell_num,
                              vertical=vertical)
        sheet_datas['Среднее значение потребления в процессе сборки, mA'](worksheet, report.voltage_average, 9,
                                                                          elem_num, merge=True,
                                                                          merge_num=merge_cell_num, vertical=vertical)

    merge_voltage_cell_num = metric_report.net_compile_amount * metric_report.draw_imgs_amount
    common_voltage_average = metric_report.voltage_average
    for elem_num, report in enumerate(draw_reports, start_position):
        draw_report_counter += 1
        sheet_datas[' '](worksheet, f'http://{server_ip}/drawed/{report.pk}', f'Отчет об отрисовке #{draw_report_counter}({report.pk})',
                         10, elem_num, vertical=vertical)
        sheet_datas['Графики этапа отрисовки ценников'](worksheet, get_grafana_url(report), f'смотреть',
                          11, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 50%'](worksheet, report.p50, 12, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 75%'](worksheet, report.p75, 13, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 90%'](worksheet, report.p90, 14, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 95%'](worksheet, report.p95, 15, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 96%'](worksheet, report.p96, 16, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 97%'](worksheet, report.p97, 17, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 98%'](worksheet, report.p98, 18, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 99%'](worksheet, report.p99, 19, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 99.5%'](worksheet, report.p995, 20, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 99.9%'](worksheet, report.p999, 21, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 100%'](worksheet, report.p100, 22, elem_num, vertical=vertical)
        sheet_datas['Не отрисовано, шт '](worksheet, report.not_drawed_esl, 23, elem_num, vertical=vertical)
        sheet_datas['Среднее значение потребления в процессе отрисовки, mA'](worksheet, report.voltage_average, 24,
                                                                             elem_num, vertical=vertical)
        sheet_datas['Средний размер изображений ценников'](worksheet, report.images_size_average, 25,
                                                                             elem_num, vertical=vertical)
        sheet_datas['Среднее значение потребления (общее), mA'](worksheet, common_voltage_average, 26, start_position,
                                                                merge=True, merge_num=merge_voltage_cell_num,
                                                                vertical=vertical)

    return workbook


def create_common_expanded_sheet(workbook, metric_report, vertical=False, start_position=1):
    net_reports = NetCompileReport.objects.filter(metric_report=metric_report).order_by('create_date_time')
    draw_reports = DrawImgsReport.objects.filter(metric_report=metric_report).order_by('create_date_time')
    worksheet = workbook.create_sheet('Сводный отчет (расширенный)', 0)

    sheet_datas = {'Дата и время начала сборки': partial(add_data_to_cell),
                   'Время сборки сети': partial(add_data_to_cell),
                   'Процент сборки сети': partial(add_data_to_cell),
                   'Статус': partial(add_data_to_cell),
                   'Время сборки 10%': partial(add_data_to_cell),
                   'Время сборки 20%': partial(add_data_to_cell),
                   'Время сборки 30%': partial(add_data_to_cell),
                   'Время сборки 40%': partial(add_data_to_cell),
                   'Время сборки 50%': partial(add_data_to_cell),
                   'Время сборки 75%': partial(add_data_to_cell),
                   'Время сборки 90%': partial(add_data_to_cell),
                   'Время сборки 95%': partial(add_data_to_cell),
                   'Время сборки 96%': partial(add_data_to_cell),
                   'Время сборки 97%': partial(add_data_to_cell),
                   'Время сборки 98%': partial(add_data_to_cell),
                   'Время сборки 99%': partial(add_data_to_cell),
                   'Время сборки 99.5%': partial(add_data_to_cell),
                   'Время сборки 99.9%': partial(add_data_to_cell),
                   'Время сборки 100%': partial(add_data_to_cell),
                   'Собранность сети за 10 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 20 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 30 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 40 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 50 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 60 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 70 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 80 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 90 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 100 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 110 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 120 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 130 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 140 мин., %': partial(add_data_to_cell),
                   'Собранность сети за 150 мин., %': partial(add_data_to_cell),
                   'Фактическое количество ESL': partial(add_data_to_cell),
                   'Сборка считается успешной при, %': partial(add_data_to_cell),
                   'Среднее значение потребления в процессе сборки, mA': partial(add_data_to_cell),
                   'Дата и время создания отчета': partial(add_data_to_cell),
                   'Дата и время окончания отчета': partial(add_data_to_cell),
                   'Предельное время отрисовки, мин': partial(add_data_to_cell),
                   'Время отрисовки 10%': partial(add_data_to_cell),
                   'Время отрисовки 20%': partial(add_data_to_cell),
                   'Время отрисовки 30%': partial(add_data_to_cell),
                   'Время отрисовки 40%': partial(add_data_to_cell),
                   'Время отрисовки 50%': partial(add_data_to_cell),
                   'Время отрисовки 75%': partial(add_data_to_cell),
                   'Время отрисовки 90%': partial(add_data_to_cell),
                   'Время отрисовки 95%': partial(add_data_to_cell),
                   'Время отрисовки 96%': partial(add_data_to_cell),
                   'Время отрисовки 97%': partial(add_data_to_cell),
                   'Время отрисовки 98%': partial(add_data_to_cell),
                   'Время отрисовки 99%': partial(add_data_to_cell),
                   'Время отрисовки 99.5%': partial(add_data_to_cell),
                   'Время отрисовки 99.9%': partial(add_data_to_cell),
                   'Время отрисовки 100%': partial(add_data_to_cell),
                   'Отрисовано за 10 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 20 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 30 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 40 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 50 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 60 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 70 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 80 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 90 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 100 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 110 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 120 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 130 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 140 мин., %': partial(add_data_to_cell),
                   'Отрисовано за 150 мин., %': partial(add_data_to_cell),
                   'Не отрисовано, шт ': partial(add_data_to_cell),
                   'Отрисовано ценников, шт': partial(add_data_to_cell),
                   'Процент отрисовки': partial(add_data_to_cell),
                   'Тип отрисовки ценников': partial(add_data_to_cell),
                   'Статус ': partial(add_data_to_cell),
                   'Среднее значение потребления в процессе отрисовки, mA': partial(add_data_to_cell),
                   }

    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    # Расширение колонок
    expands_columns_width(worksheet)

    # заполнение таблицы даными
    start_position += 1

    merge_cell_num = metric_report.draw_imgs_amount
    for elem_num, report in step_enumerate(net_reports, start_position, merge_cell_num):
        sheet_datas['Дата и время начала сборки'](worksheet, report.create_date_time, 1, elem_num, merge=True,
                                                  merge_num=merge_cell_num,
                                                  vertical=vertical)
        sheet_datas['Время сборки сети'](worksheet, report.elapsed_time, 2, elem_num, merge=True,
                                         merge_num=merge_cell_num,
                                         vertical=vertical)
        sheet_datas['Процент сборки сети'](worksheet, report.final_percent, 3, elem_num, merge=True,
                                           merge_num=merge_cell_num,
                                           vertical=vertical)
        sheet_datas['Статус'](worksheet, report.status, 4, elem_num, merge=True, merge_num=merge_cell_num,
                              vertical=vertical)
        sheet_datas['Время сборки 10%'](worksheet, report.p10, 5, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 20%'](worksheet, report.p20, 6, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 30%'](worksheet, report.p30, 7, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 40%'](worksheet, report.p40, 8, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 50%'](worksheet, report.p50, 9, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 75%'](worksheet, report.p75, 10, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 90%'](worksheet, report.p90, 11, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 95%'](worksheet, report.p95, 12, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 96%'](worksheet, report.p96, 13, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 97%'](worksheet, report.p97, 14, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 98%'](worksheet, report.p98, 15, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 99%'](worksheet, report.p99, 16, elem_num, merge=True, merge_num=merge_cell_num,
                                        vertical=vertical)
        sheet_datas['Время сборки 99.5%'](worksheet, report.p995, 17, elem_num, merge=True, merge_num=merge_cell_num,
                                          vertical=vertical)
        sheet_datas['Время сборки 99.9%'](worksheet, report.p999, 18, elem_num, merge=True, merge_num=merge_cell_num,
                                          vertical=vertical)
        sheet_datas['Время сборки 100%'](worksheet, report.p100, 19, elem_num, merge=True, merge_num=merge_cell_num,
                                         vertical=vertical)
        sheet_datas['Собранность сети за 10 мин., %'](worksheet, report.t10, 20, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 20 мин., %'](worksheet, report.t20, 21, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 30 мин., %'](worksheet, report.t30, 22, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 40 мин., %'](worksheet, report.t40, 23, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 50 мин., %'](worksheet, report.t50, 24, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 60 мин., %'](worksheet, report.t60, 25, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 70 мин., %'](worksheet, report.t70, 26, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 80 мин., %'](worksheet, report.t80, 27, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 90 мин., %'](worksheet, report.t90, 28, elem_num, merge=True,
                                                      merge_num=merge_cell_num,
                                                      vertical=vertical)
        sheet_datas['Собранность сети за 100 мин., %'](worksheet, report.t100, 29, elem_num, merge=True,
                                                       merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Собранность сети за 110 мин., %'](worksheet, report.t110, 30, elem_num, merge=True,
                                                       merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Собранность сети за 120 мин., %'](worksheet, report.t120, 31, elem_num, merge=True,
                                                       merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Собранность сети за 130 мин., %'](worksheet, report.t130, 32, elem_num, merge=True,
                                                       merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Собранность сети за 140 мин., %'](worksheet, report.t140, 33, elem_num, merge=True,
                                                       merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Собранность сети за 150 мин., %'](worksheet, report.t150, 34, elem_num, merge=True,
                                                       merge_num=merge_cell_num, vertical=vertical)
        sheet_datas['Фактическое количество ESL'](worksheet, report.fact_total_esl, 35, elem_num, merge=True,
                                                  merge_num=merge_cell_num,
                                                  vertical=vertical)
        sheet_datas['Сборка считается успешной при, %'](worksheet, report.success_percent, 36, elem_num, merge=True,
                                                        merge_num=merge_cell_num,
                                                        vertical=vertical)
        sheet_datas['Среднее значение потребления в процессе сборки, mA'](worksheet, report.voltage_average, 37,
                                                                          elem_num,
                                                                          merge=True, merge_num=merge_cell_num,
                                                                          vertical=vertical)

    for elem_num, report in enumerate(draw_reports, start_position):
        sheet_datas['Дата и время создания отчета'](worksheet, report.create_date_time, 38, elem_num, vertical=vertical)
        sheet_datas['Дата и время окончания отчета'](worksheet, report.date_time_finish, 39, elem_num,
                                                     vertical=vertical)
        sheet_datas['Предельное время отрисовки, мин'](worksheet, report.draw_imgs_limit_mins, 40, elem_num,
                                                       vertical=vertical)
        sheet_datas['Время отрисовки 10%'](worksheet, report.p10, 41, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 20%'](worksheet, report.p20, 42, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 30%'](worksheet, report.p30, 43, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 40%'](worksheet, report.p40, 44, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 50%'](worksheet, report.p50, 45, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 75%'](worksheet, report.p75, 46, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 90%'](worksheet, report.p90, 47, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 95%'](worksheet, report.p95, 48, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 96%'](worksheet, report.p96, 49, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 97%'](worksheet, report.p97, 50, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 98%'](worksheet, report.p98, 51, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 99%'](worksheet, report.p99, 52, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 99.5%'](worksheet, report.p995, 53, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 99.9%'](worksheet, report.p999, 54, elem_num, vertical=vertical)
        sheet_datas['Время отрисовки 100%'](worksheet, report.p100, 55, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 10 мин., %'](worksheet, report.t10, 56, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 20 мин., %'](worksheet, report.t20, 57, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 30 мин., %'](worksheet, report.t30, 58, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 40 мин., %'](worksheet, report.t40, 59, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 50 мин., %'](worksheet, report.t50, 60, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 60 мин., %'](worksheet, report.t60, 61, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 70 мин., %'](worksheet, report.t70, 62, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 80 мин., %'](worksheet, report.t80, 63, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 90 мин., %'](worksheet, report.t90, 64, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 100 мин., %'](worksheet, report.t100, 65, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 110 мин., %'](worksheet, report.t110, 66, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 120 мин., %'](worksheet, report.t120, 67, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 130 мин., %'](worksheet, report.t130, 68, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 140 мин., %'](worksheet, report.t140, 69, elem_num, vertical=vertical)
        sheet_datas['Отрисовано за 150 мин., %'](worksheet, report.t150, 70, elem_num, vertical=vertical)
        sheet_datas['Не отрисовано, шт '](worksheet, report.not_drawed_esl, 71, elem_num, vertical=vertical)
        sheet_datas['Отрисовано ценников, шт'](worksheet, report.drawed_esl, 72, elem_num, vertical=vertical)
        sheet_datas['Процент отрисовки'](worksheet, report.final_percent, 73, elem_num, vertical=vertical)
        sheet_datas['Тип отрисовки ценников'](worksheet, report.draw_imgs_type, 74, elem_num, vertical=vertical)
        sheet_datas['Статус '](worksheet, report.status, 75, elem_num, vertical=vertical)
        sheet_datas['Среднее значение потребления в процессе отрисовки, mA'](worksheet, report.voltage_average, 76,
                                                                             elem_num, vertical=vertical)
    return workbook


def create_configuration_sheet(workbook, report, vertical=True, start_position=1):
    worksheet = workbook.create_sheet('Конфигурация', 0)

    sheet_datas = {'Номера щитов': partial(add_data_to_cell),
                   'Конфигурация системы': partial(add_data_to_cell),
                   'Количество ценников стенда, шт': partial(add_data_to_cell),
                   'Количество РУ, шт': partial(add_data_to_cell),
                   'Конфигурация РУ': partial(add_data_to_cell),
                   'Количество донглов на РУ, шт': partial(add_data_to_cell),
                   'Версия СУМ': partial(add_data_to_cell),
                   'Версия Хаоса': partial(add_data_to_cell),
                   'Число этажей дерева': partial(add_data_to_cell),
                   'Версия драйвера': partial(add_data_to_cell),
                   'Версия прошивки ЭЦ': partial(add_data_to_cell),
                   'HW версия ЭЦ': partial(add_data_to_cell),
                   'HW версия донглов': partial(add_data_to_cell),
                   }

    # создание шапки таблицы
    insert_table_titles(worksheet, sheet_datas, vertical=vertical, bolt=True)

    # заполнение таблицы даными
    start_position += 1
    sheet_datas['Номера щитов'](worksheet, report.shields_num, 1, start_position, vertical=vertical)
    sheet_datas['Конфигурация системы'](worksheet, report.hardware_config, 2, start_position, vertical=vertical)
    sheet_datas['Количество ценников стенда, шт'](worksheet, report.total_esl, 3, start_position, vertical=vertical)
    sheet_datas['Количество РУ, шт'](worksheet, report.dd_nums, 4, start_position, vertical=vertical)
    sheet_datas['Конфигурация РУ'](worksheet, report.dd_configuration, 5, start_position, vertical=vertical)
    sheet_datas['Количество донглов на РУ, шт'](worksheet, report.dd_dongles_num, 6, start_position, vertical=vertical)
    sheet_datas['Версия СУМ'](worksheet, report.version_sum, 7, start_position, vertical=vertical)
    sheet_datas['Версия Хаоса'](worksheet, report.version_chaos, 8, start_position, vertical=vertical)
    sheet_datas['Число этажей дерева'](worksheet, report.tree_floor_num, 9, start_position, vertical=vertical)
    sheet_datas['Версия драйвера'](worksheet, report.version_driver, 10, start_position, vertical=vertical)
    sheet_datas['Версия прошивки ЭЦ'](worksheet, report.version_esl_firmware, 11, start_position, vertical=vertical)
    sheet_datas['HW версия ЭЦ'](worksheet, report.version_esl_hw, 12, start_position, vertical=vertical)
    sheet_datas['HW версия донглов'](worksheet, report.version_dongles_hw, 13, start_position, vertical=vertical)

    # Расширение колонок
    expands_columns_width(worksheet)
    return workbook


if __name__ == '__main__':
    pass

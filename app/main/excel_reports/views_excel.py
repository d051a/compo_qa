from main.models import Statistic,  MetricReport, Chaos, DrawImgsReport, NetCompileReport, Configuration
from main.excel_reports.excel_tools import create_excel_cheet, create_draw_imgs_stats_sheet,\
    create_net_compile_stats_sheet, create_all_statistics_sheet, create_common_sheet, create_common_expanded_sheet,\
    create_configuration_sheet, create_net_compile_sheet, create_draw_imgs_sheet
from django.http import HttpResponse
from datetime import datetime
from django.utils import timezone
from django.contrib.sites.shortcuts import get_current_site
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import Workbook
from main.excel_reports.excel_report_fields import net_compile_short_report_draw_fields,\
    draw_imgs_stats_short_report_draw_fields, chaos_configuration_fields


def metric_report_export_to_xlsx(request, metric_report_id):
    """
    Генерация excel-отчета со страницы снятия общих метрик /metrics/<report_ID>/
    :param request:
    :param metric_report_id:
    :return:
    """
    metrics_report = MetricReport.objects.get(pk=metric_report_id)
    net_compile_reports = NetCompileReport.objects.filter(metric_report=metrics_report)
    draw_imgs_reports = DrawImgsReport.objects.filter(metric_report=metrics_report)
    report_start_time = metrics_report.create_date_time
    report_finish_time = metrics_report.date_time_finish
    server_ip = get_current_site(request).domain
    if report_finish_time is None:
        report_finish_time = timezone.localtime()
    all_statistics_report = Statistic.objects.filter(metric_report=metrics_report).filter(
        date_time__range=(report_start_time, report_finish_time)).order_by('-date_time')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-metric_report_{chaos_name}.xlsx'.format(
        date=datetime.now().strftime('%d.%m.%Y_%H.%M.%S'),
        chaos_name=metrics_report.chaos.name
    )
    workbook = Workbook()

    for report in draw_imgs_reports:
        workbook = create_draw_imgs_stats_sheet(workbook, report, vertical=False)
    for report in net_compile_reports:
        workbook = create_net_compile_stats_sheet(workbook, report, vertical=False)

    workbook = create_all_statistics_sheet(workbook, all_statistics_report, vertical=False)
    workbook = create_common_expanded_sheet(workbook, metrics_report,vertical=True)
    workbook = create_common_sheet(workbook, metrics_report, server_ip, vertical=True)
    try:
        configuration = Configuration.objects.get(metric_report=metrics_report)
        workbook = create_configuration_sheet(workbook, configuration, chaos_configuration_fields)
    except ObjectDoesNotExist:
        pass
    workbook.save(response)
    return response


def chaos_stats_export_to_xlsx(request, chaos_id):
    """
    Генерация excel-отчета со страницы (chaoses/<chaos_ID>/stats) по всем сборкам\отрисовкам
    :param request:
    :param chaos_id:
    :return:
    """
    chaos = Chaos.objects.get(pk=chaos_id)
    net_compile_reports = NetCompileReport.objects.filter(chaos=chaos).order_by('-create_date_time')
    draw_imgs_reports = DrawImgsReport.objects.filter(chaos=chaos).order_by('-create_date_time')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-nets_draws_report_{chaos_name}.xlsx'.format(
        date=datetime.now().strftime('%d.%m.%Y_%H.%M.%S'),
        chaos_name=chaos.name
    )

    workbook = Workbook()
    workbook = create_excel_cheet(workbook, net_compile_reports, net_compile_short_report_draw_fields)
    workbook = create_excel_cheet(workbook, draw_imgs_reports, draw_imgs_stats_short_report_draw_fields)
    workbook.save(response)
    return response


def draw_imgs_report_export_to_xlsx(request, draw_imgs_report_id):
    """
    Генерация excel-отчета со страницы (drawed/<draw_imgs_report_id>) отрисовки ценников
    :param request:
    :param draw_imgs_report_id:
    :return:
    """
    current_time = timezone.localtime()
    draw_imgs_report = DrawImgsReport.objects.get(pk=draw_imgs_report_id)
    chaos = draw_imgs_report.chaos
    draw_imgs_report_start_time = draw_imgs_report.create_date_time
    draw_imgs_report_end_time = draw_imgs_report.date_time_finish
    if draw_imgs_report_end_time is None:
        draw_imgs_report_end_time = current_time
    all_statistics_report = Statistic.objects.filter(chaos=chaos).\
        filter(date_time__range=[draw_imgs_report_start_time, draw_imgs_report_end_time]).order_by('date_time')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-draw_imgs_report_{chaos_name}.xlsx'.format(
        date=datetime.now().strftime('%d.%m.%Y_%H.%M.%S'),
        chaos_name=chaos.name
    )

    workbook = Workbook()
    workbook = create_draw_imgs_stats_sheet(workbook, draw_imgs_report, vertical=False)
    workbook = create_draw_imgs_sheet(workbook, draw_imgs_report, vertical=True)
    workbook = create_all_statistics_sheet(workbook, all_statistics_report, vertical=False)
    try:
        configuration = Configuration.objects.get(drawimgs_report=draw_imgs_report)
        workbook = create_configuration_sheet(workbook, configuration, chaos_configuration_fields)
    except ObjectDoesNotExist:
        pass
    workbook.save(response)
    return response


def net_compile_report_export_to_xlsx(request, net_compile_id):
    """
    Генерация excel-отчета со страницы (netcompiles/<net_compile_id>) сборки сети
    :param request:
    :param net_compile_id:
    :return:
    """
    current_time = timezone.localtime()
    net_compile_report = NetCompileReport.objects.get(pk=net_compile_id)
    chaos = net_compile_report.chaos
    net_compile_start_time = net_compile_report.create_date_time
    net_compile_end_time = net_compile_report.date_time_finish
    if net_compile_end_time is None:
        net_compile_end_time = current_time
    all_statistics_report = Statistic.objects.filter(chaos=chaos).\
        filter(date_time__range=[net_compile_start_time, net_compile_end_time]).order_by('date_time')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-net_compile_report_{chaos_name}.xlsx'.format(
        date=datetime.now().strftime('%d.%m.%Y_%H.%M.%S'),
        chaos_name=chaos.name
    )

    workbook = Workbook()
    workbook = create_net_compile_stats_sheet(workbook, net_compile_report, vertical=False)
    workbook = create_net_compile_sheet(workbook, net_compile_report, vertical=True)
    workbook = create_all_statistics_sheet(workbook, all_statistics_report, vertical=False)
    try:
        configuration = Configuration.objects.get(netcompile_report=net_compile_report)
        workbook = create_configuration_sheet(workbook, configuration, chaos_configuration_fields)
    except ObjectDoesNotExist:
        pass
    workbook.save(response)
    return response
